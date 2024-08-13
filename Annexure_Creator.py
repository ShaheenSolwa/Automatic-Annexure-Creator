import pandas as pd
import streamlit as st
from pandas.api.types import (
    is_categorical_dtype,
    is_datetime64_any_dtype,
    is_numeric_dtype,
    is_object_dtype,
)
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
import time

st.set_page_config(layout="wide")

def filter_dataframe(df: pd.DataFrame, key: str) -> pd.DataFrame:
    """
    Adds a UI on top of a dataframe to let viewers filter columns

    Args:
        df (pd.DataFrame): Original dataframe

    Returns:
        pd.DataFrame: Filtered dataframe
    """
    modify = st.checkbox("Add filters", key=key)

    if not modify:
        return df

    df = df.copy()

    # Try to convert datetimes into a standard format (datetime, no timezone)
    for col in df.columns:
        if is_object_dtype(df[col]):
            try:
                df[col] = pd.to_datetime(df[col])
            except Exception:
                pass

        if is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.tz_localize(None)

    modification_container = st.container()

    with modification_container:
        to_filter_columns = st.multiselect("Filter dataframe on", df.columns)
        for column in to_filter_columns:
            left, right = st.columns((1, 20))
            left.write("↳")
            # Treat columns with < 10 unique values as categorical
            if is_categorical_dtype(df[column]) or df[column].nunique() < 10:
                default_values = df[column].unique().tolist()
                user_cat_input = right.multiselect(f"Values for {column}", default_values, default=default_values, key=f"{key}_{column}")
                df = df[df[column].isin(user_cat_input)]
            elif is_numeric_dtype(df[column]):
                _min = float(df[column].min())
                _max = float(df[column].max())
                step = (_max - _min) / 100
                user_num_input = right.slider(
                    f"Values for {column}",
                    _min,
                    _max,
                    (_min, _max),
                    step=step,
                    key = f"{key}_{column}"
                )
                df = df[df[column].between(*user_num_input)]
            elif is_datetime64_any_dtype(df[column]):
                user_date_input = right.date_input(
                    f"Values for {column}",
                    value=(
                        df[column].min(),
                        df[column].max(),
                    ),
                    key=f"{key}_{column}"
                )
                if len(user_date_input) == 2:
                    user_date_input = tuple(map(pd.to_datetime, user_date_input))
                    start_date, end_date = user_date_input
                    df = df.loc[df[column].between(start_date, end_date)]
            else:
                user_text_input = right.text_input(
                    f"Substring or regex in {column}",
                    key=f"{key}_{column}"
                )
                if user_text_input:
                    df = df[df[column].str.contains(user_text_input)]

    return df


db_or_excel = st.selectbox("Choose a data entry point", options=["", "Database", "Excel/ CSV"])
if db_or_excel != "":
    if db_or_excel == "Database":
        pass
    elif db_or_excel == "Excel/ CSV":
        multiple_files = st.selectbox("", options=["", "Single", "Multiple"])
        if multiple_files != "":
            if multiple_files == 'Single':
                file = st.file_uploader("Upload your file", type=['xlsx', 'csv'])
                if file is not None:
                    extension = file.name.split('.')[-1]
                    df = pd.DataFrame()
                    if extension == 'csv':
                        df = pd.read_csv(file)

                        filtered_df = filter_dataframe(df)
                        st.dataframe(filtered_df, use_container_width=True)

                        annexure_name = st.text_input("Enter a name for the Annexure", key='annex5')
                        if st.button("Create Annexure"):
                            try:
                                filtered_df.to_excel(f"{annexure_name}.xlsx", index=False)
                                # Load the workbook
                                workbook = openpyxl.load_workbook(f"{annexure_name}.xlsx")

                                # Select the first sheet
                                sheet = workbook.active

                                # Select the first row
                                row = sheet[1]

                                # Bold the text
                                for cell in row:
                                    cell.font = Font(size=11, color='00000000', bold=True)
                                    cell.fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')

                                for col in sheet.columns:
                                    max_length = 0
                                    column = col[0].column_letter  # Get the column name
                                    for cell in col:
                                        try:  # Necessary to avoid error on empty cells
                                            if len(str(cell.value)) > max_length:
                                                max_length = len(str(cell.value))
                                        except:
                                            pass
                                    adjusted_width = (max_length + 2) * 1.2
                                    sheet.column_dimensions[column].width = adjusted_width

                                for row in sheet.iter_rows():
                                    for cell in row:
                                        # Set the border style for the cell
                                        cell.border = Border(
                                            left=openpyxl.styles.Side(style='thin'),
                                            right=openpyxl.styles.Side(style='thin'),
                                            top=openpyxl.styles.Side(style='thin'),
                                            bottom=openpyxl.styles.Side(style='thin')
                                        )

                                # Save the changes
                                workbook.save(f"{annexure_name}.xlsx")

                                st.success("Successfully created the Annexure!")
                                time.sleep(5)

                            except Exception as e:
                                st.warning("Failed to create Annexure!")
                                st.warning(str(e))
                                time.sleep(5)


                    else:
                        xlsx = pd.ExcelFile(file)
                        sheet_names = xlsx.sheet_names
                        selected_sheet = st.selectbox("Select a sheet to display", sheet_names)
                        if selected_sheet is not None:
                            df = pd.read_excel(file, sheet_name=selected_sheet)

                            filtered_df = filter_dataframe(df)
                            st.dataframe(filtered_df, use_container_width=True)

                            annexure_name = st.text_input("Enter a name for the Annexure", key='anex2')

                            if st.button("Create Annexure"):
                                try:
                                    filtered_df.to_excel(f"{annexure_name}.xlsx", index=False)
                                    # Load the workbook
                                    workbook = openpyxl.load_workbook(f"{annexure_name}.xlsx")

                                    # Select the first sheet
                                    sheet = workbook.active

                                    # Select the first row
                                    row = sheet[1]

                                    # Bold the text
                                    for cell in row:
                                        cell.font = Font(size=11, color='00000000', bold=True)
                                        cell.fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')

                                    for col in sheet.columns:
                                        max_length = 0
                                        column = col[0].column_letter  # Get the column name
                                        for cell in col:
                                            try:  # Necessary to avoid error on empty cells
                                                if len(str(cell.value)) > max_length:
                                                    max_length = len(str(cell.value))
                                            except:
                                                pass
                                        adjusted_width = (max_length + 2) * 1.2
                                        sheet.column_dimensions[column].width = adjusted_width

                                    border = Border(left=Side(style='thin'),
                                                    right=Side(style='thin'),
                                                    top=Side(style='thin'),
                                                    bottom=Side(style='thin'))

                                    # Set the border for a range of cells
                                    for row in range(1, 6):
                                        for col in range(1, 4):
                                            cell = sheet.cell(row=row, column=col)
                                            cell.border = border

                                    # Save the changes
                                    workbook.save(f"{annexure_name}.xlsx")

                                    st.success("Successfully created the Annexure!")
                                    time.sleep(5)

                                except Exception as e:
                                    st.warning("Failed to create Annexure!")
                                    st.warning(str(e))
                                    time.sleep(5)
    
            elif multiple_files == 'Multiple':
                files = st.file_uploader("Upload your file", type=['xlsx', 'csv'], accept_multiple_files=True)
                if len(files) > 0 and len(files) <= 2:
                    col1, col2 = st.columns(2)
                    with col1:
                        file = files[0]
                        if file.name.split('.')[-1] == 'csv':
                            df = pd.read_csv(file)

                            filtered_df = filter_dataframe(df, key='cb1')
                            st.dataframe(filtered_df, use_container_width=True)


                        else:
                            xlsx = pd.ExcelFile(file)
                            sheet_names = xlsx.sheet_names
                            selected_sheet1 = st.selectbox("Select a sheet to display", sheet_names, key='ss1')
                            if selected_sheet1 is not None:
                                df = pd.read_excel(file, sheet_name=selected_sheet1)
                                filtered_df = filter_dataframe(df, key='cb1')
                                st.dataframe(filtered_df, use_container_width=True)

                        annexure_name = st.text_input("Enter a name for the annexure", key="anex3")

                        if st.button("Create Annexure", key='btn1'):
                            try:
                                filtered_df.to_excel(f"{annexure_name}.xlsx", index=False)
                                # Load the workbook
                                workbook = openpyxl.load_workbook(f"{annexure_name}.xlsx")

                                # Select the first sheet
                                sheet = workbook.active

                                # Select the first row
                                row = sheet[1]

                                # Bold the text
                                for cell in row:
                                    cell.font = Font(size=11, color='00000000', bold=True)
                                    cell.fill = PatternFill(start_color='808080', end_color='808080',
                                                            fill_type='solid')

                                for col in sheet.columns:
                                    max_length = 0
                                    column = col[0].column_letter  # Get the column name
                                    for cell in col:
                                        try:  # Necessary to avoid error on empty cells
                                            if len(str(cell.value)) > max_length:
                                                max_length = len(str(cell.value))
                                        except:
                                            pass
                                    adjusted_width = (max_length + 2) * 1.2
                                    sheet.column_dimensions[column].width = adjusted_width

                                for row in sheet.iter_rows():
                                    for cell in row:
                                        # Set the border style for the cell
                                        cell.border = Border(
                                            left=openpyxl.styles.Side(style='thin'),
                                            right=openpyxl.styles.Side(style='thin'),
                                            top=openpyxl.styles.Side(style='thin'),
                                            bottom=openpyxl.styles.Side(style='thin')
                                        )

                                # Save the changes
                                workbook.save(f"{annexure_name}.xlsx")

                                st.success("Successfully created the Annexure!")
                                time.sleep(5)

                            except Exception as e:
                                st.warning("Failed to create Annexure!")
                                st.warning(str(e))
                                time.sleep(5)

                    with col2:
                        file = files[1]
                        if file.name.split('.')[-1] == 'csv':
                            df = pd.read_csv(file)
                            filtered_df = filter_dataframe(df, key='cb2')
                            st.dataframe(filtered_df, use_container_width=True)
                        else:
                            xlsx = pd.ExcelFile(file)
                            sheet_names = xlsx.sheet_names
                            selected_sheet1 = st.selectbox("Select a sheet to display", sheet_names, key='ss2')
                            if selected_sheet1 is not None:
                                df = pd.read_excel(file, sheet_name=selected_sheet1)
                                filtered_df = filter_dataframe(df, key='cb2')
                                st.dataframe(filtered_df, use_container_width=True)

                        annexure_name = st.text_input("Enter a name for the annexure", key="anex4")

                        if st.button("Create Annexure", key='btn2'):
                            try:
                                filtered_df.to_excel(f"{annexure_name}.xlsx", index=False)
                                # Load the workbook
                                workbook = openpyxl.load_workbook(f"{annexure_name}.xlsx")

                                # Select the first sheet
                                sheet = workbook.active

                                # Select the first row
                                row = sheet[1]

                                # Bold the text
                                for cell in row:
                                    cell.font = Font(size=11, color='00000000', bold=True)
                                    cell.fill = PatternFill(start_color='808080', end_color='808080',
                                                            fill_type='solid')

                                for col in sheet.columns:
                                    max_length = 0
                                    column = col[0].column_letter  # Get the column name
                                    for cell in col:
                                        try:  # Necessary to avoid error on empty cells
                                            if len(str(cell.value)) > max_length:
                                                max_length = len(str(cell.value))
                                        except:
                                            pass
                                    adjusted_width = (max_length + 2) * 1.2
                                    sheet.column_dimensions[column].width = adjusted_width

                                for row in sheet.iter_rows():
                                    for cell in row:
                                        # Set the border style for the cell
                                        cell.border = Border(
                                            left=openpyxl.styles.Side(style='thin'),
                                            right=openpyxl.styles.Side(style='thin'),
                                            top=openpyxl.styles.Side(style='thin'),
                                            bottom=openpyxl.styles.Side(style='thin')
                                        )

                                # Save the changes
                                workbook.save(f"{annexure_name}.xlsx")

                                st.success("Successfully created the Annexure!")
                                time.sleep(5)

                            except Exception as e:
                                st.warning("Failed to create Annexure!")
                                st.warning(str(e))
                                time.sleep(5)
